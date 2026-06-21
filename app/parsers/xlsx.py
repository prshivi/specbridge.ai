import io
import zipfile
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.exceptions import DocumentParsingError
from app.models.document import BlockType, DocumentBlock, NormalizedDocument
from app.parsers.base import DocumentParser, normalize_text


def format_cell(value: object) -> str:
    """Convert a spreadsheet cell value to stable text."""
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


class XlsxParser(DocumentParser):
    """Extract worksheet values from XLSX workbooks."""

    def parse(self, content: bytes) -> str:
        return self.parse_document(content).text

    def parse_document(self, content: bytes) -> NormalizedDocument:
        try:
            workbook = load_workbook(
                filename=io.BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except (
            InvalidFileException,
            ValueError,
            KeyError,
            OSError,
            zipfile.BadZipFile,
        ) as error:
            raise DocumentParsingError("The XLSX file could not be parsed.") from error

        sections: list[str] = []
        blocks: list[DocumentBlock] = []
        try:
            for worksheet in workbook.worksheets:
                rows = []
                for row in worksheet.iter_rows(values_only=True):
                    values = [format_cell(value) for value in row]
                    while values and not values[-1]:
                        values.pop()
                    if any(values):
                        rows.append("\t".join(values))
                if rows:
                    value = "\n".join(rows)
                    sections.append(f"# Sheet: {worksheet.title}\n{value}")
                    blocks.append(
                        DocumentBlock(
                            block_type=BlockType.HEADING,
                            text=worksheet.title,
                            heading=worksheet.title,
                            heading_level=1,
                        )
                    )
                    blocks.append(
                        DocumentBlock(
                            block_type=BlockType.TABLE,
                            text=value,
                            heading=worksheet.title,
                        )
                    )
        finally:
            workbook.close()

        return NormalizedDocument(
            text=normalize_text("\n\n".join(sections)),
            blocks=blocks,
        )
